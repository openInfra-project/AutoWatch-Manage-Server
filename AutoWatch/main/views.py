from AutoWatch import settings
from django.db.models.fields import NullBooleanField
from django.db.models.query import QuerySet
from django.http import response
from django.shortcuts import redirect, render
from django.urls.conf import path
from home.models import User
from .models import Analytics, Room
import string
import random
from django.views.generic import ListView
from collections import OrderedDict
from .fusioncharts import FusionCharts
import base64

import simplejson
from django.core.files.storage import FileSystemStorage
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.http.response import HttpResponse, JsonResponse
from django.core import serializers

import os
import base64
from luxand import luxand
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader


def main(request):
    res_data = {}
    user_session = request.session.get('user')              # 로그인 체크
    fs = FileSystemStorage()
    if user_session:
        user = User.objects.get(pk=user_session)
        res_data['username'] = user.username                # mypage 정보
        res_data['email'] = user.email
        res_data['register'] = user.registerd_date
        res_data['userimg'] = fs.url(user.image)

        if res_data['userimg'] == "/media/":               # 이미지 체크
            res_data['img_check'] = 0                      # 이미지 널
        else:
            res_data['img_check'] = 1

        if request.method == 'GET':
            return render(request, 'main.html', res_data)
        elif request.method == 'POST':
            userimage = request.FILES['user-img-change']
            res_data['userimg'] = fs.url(userimage)
            user.image = userimage
            user.save()
            res_data['img_check'] = 1
            return render(request, 'main.html', res_data)
    else:
        return redirect('/login')


def makeroom(request):
    res_data = {}
    user_session = request.session.get('user')
    fs = FileSystemStorage()
    if user_session:
        user = User.objects.get(pk=user_session)            # 로그인 체크
        res_data['username'] = user.username                # mypage 정보
        res_data['email'] = user.email
        res_data['register'] = user.registerd_date
        res_data['string'] = ''.join(random.choice(
            string.ascii_uppercase + string.digits)for _ in range(7))  # 랜덤 문자열 생성
        res_data['userimg'] = fs.url(user.image)

        if res_data['userimg'] == "/media/":               # 이미지가 있는지 체크
            res_data['img_check'] = 0
        else:
            res_data['img_check'] = 1

        if request.method == 'GET':
            return render(request, 'makeroom.html', res_data)
        elif request.method == 'POST':
            post_type = request.POST.get('user-img-btn')
            if post_type == "모달":  # 버튼 값을 읽어 POST 구분, 이미지 변경 POST
                # 이미지 변경 저장
                userimage = request.FILES['user-img-change']
                res_data['userimg'] = fs.url(userimage)
                res_data['img_check'] = 1
                user.image = userimage
                user.save()
                return render(request, 'makeroom.html', res_data)
            else:  # Room create POST
                # makeroom POST 값
                room_name = request.POST.get('room-name', None)
                room_password = request.POST.get('room-password', None)
                study = request.POST.getlist('study')
                exam = request.POST.getlist('exam')
                maker = user.email
                member_list = []
                if exam and not(study):
                    file = request.FILES['file']
                    # 학생명단 file
                    # 명단에서학번만 추출
                    fs = FileSystemStorage()
                    filename = fs.save(file.name, file)
                    print(filename)

                    member = load_workbook("media/" + file.name)
                    for cell in member['Sheet1']['A']:
                        member_list.append(cell.value)

                    os.remove(os.path.join(settings.MEDIA_ROOT, file.name))

                elif study and not(exam):
                    file = "NULL"

                if (len(room_name) > 7):
                    res_data['name_error'] = '방 이름을 생성해 주세요.'
                elif not(room_password):
                    res_data['password_error'] = '비밀번호를 생성해 주세요.'
                elif not(study or exam):
                    res_data['mode_error'] = 'Mode를 선택해 주세요.'
                else:
                    if (exam and not(file)):
                        res_data['mode_error'] = 'Exam Mode는 명단 첨부가 필수 입니다.'
                        # room 정보 비정상 일시
                        return render(request, 'makeroom.html', res_data)
                    else:  # 정상적으로 room 정보 기입시
                        if study and not(exam):     # mode를 db에 저장
                            mode = 'STUDY'
                        elif exam and not(study):
                            mode = 'EXAM'
                        room = Room(room_name=room_name, room_password=room_password,
                                    file=file, mode=mode, maker=maker, member_list=member_list)  # db에 room 정보 저장
                        room.save()
                        # 방을 성공적으로 만들면 room_name으로 room_session을 저장
                        request.session['room_name'] = room_name
                        return redirect('/main/makeroom/success')
                # room 정보 비정상 일시
                return render(request, 'makeroom.html', res_data)
    else:
        return redirect('/login')


def make_success(request):
    res_data = {}
    fs = FileSystemStorage()
    room_session = request.session.get(
        'room_name')   # 아까 POST 할때 session에 저장한 값 불러옴
    user_session = request.session.get('user')
    if room_session and user_session:
        user = User.objects.get(pk=user_session)    # 로그인 체크
        res_data['username'] = user.username        # mypage 정보
        res_data['email'] = user.email
        res_data['register'] = user.registerd_date

        # 가장 최근의 room_name과 session에 저장한 것을 비교함
        room = Room.objects.get(room_name=room_session)
        res_data['room_name'] = room.room_name
        res_data['room_password'] = room.room_password
        res_data['mode'] = room.mode
        res_data['maker'] = room.maker

        res_data['userimg'] = fs.url(user.image)

        if res_data['userimg'] == "/media/":               # 이미지 체크
            res_data['img_check'] = 0
        else:
            res_data['img_check'] = 1
        if request.method == 'GET':
            return render(request, 'make_success.html', res_data)
        elif request.method == 'POST':
            post_type = request.POST.get('user-img-btn')
            if post_type == "모달":  # 버튼 값을 읽어 POST 구분, 이미지 변경 POST
                # 이미지 변경 저장
                userimage = request.FILES['user-img-change']
                res_data['userimg'] = fs.url(userimage)
                res_data['img_check'] = 1
                user.image = userimage
                user.save()
                return render(request, 'make_success.html', res_data)
            else:  # Room 입장 POST
                if room.mode == "EXAM":
                    return redirect('/main/enteroom/exam1')
                elif room.mode == "STUDY":
                    return redirect('/main/enteroom/study1')
    else:
        return redirect('/login')


def enteroom(request):
    res_data = {}
    fs = FileSystemStorage()
    user_session = request.session.get('user')
    if user_session:
        user = User.objects.get(pk=user_session)    # 로그인 체크
        res_data['username'] = user.username        # mypage 정보
        res_data['email'] = user.email
        res_data['register'] = user.registerd_date
        res_data['userimg'] = fs.url(user.image)

        if res_data['userimg'] == "/media/":               # 이미지 체크
            res_data['img_check'] = 0
        else:
            res_data['img_check'] = 1

        if request.method == 'GET':
            return render(request, 'enteroom.html', res_data)
        elif request.method == 'POST':
            post_type = request.POST.get('user-img-btn')
            if post_type == "모달":  # 버튼 값을 읽어 POST 구분, 이미지 변경 POST
                # 이미지 변경 저장
                userimage = request.FILES['user-img-change']
                res_data['userimg'] = fs.url(userimage)
                res_data['img_check'] = 1
                user.image = userimage
                user.save()
                return render(request, 'enteroom.html', res_data)
            else:  # Room 입장 POST
                room_name = request.POST.get('room_name')
                room_password = request.POST.get('room_password')

                if not(room_name):
                    res_data['name_error'] = 'Room 이름을 입력하세요.'
                elif not(room_password):
                    res_data['password_error'] = 'Room 비밀번호를 입력하세요.'
                elif not(room_name and room_password):
                    res_data['all_error'] = '모든 값을 입력하세요.'
                else:
                    try:
                        # 필드명 = 값 이면 Room 객체 생성
                        room = Room.objects.get(room_name=room_name)
                    except Room.DoesNotExist:
                        # room이 없는 예외 처리
                        res_data['error'] = '존재하지 않는 Room 입니다.'
                        return render(request, 'enteroom.html', res_data)

                    # 방 입장하는 순간 room session의 기준은 입장한 방 이름
                    request.session['room_name'] = room_name
                    db_password = room.room_password
                    if db_password == room_password:     # room 정상 입장
                        if room.mode == "EXAM":
                            return redirect('/main/enteroom/exam1')
                        elif room.mode == "STUDY":
                            return redirect('/main/enteroom/study1')
                    else:
                        res_data['error'] = '비밀번호가 틀렸습니다.'
                        return render(request, 'enteroom.html', res_data)
                # room 정보 비정상 일시
                return render(request, 'enteroom.html', res_data)
    else:
        return redirect('/login')


def exam1(request):
    res_data = {}
    fs = FileSystemStorage()
    user_session = request.session.get('user')
    if user_session:
        user = User.objects.get(pk=user_session)    # 로그인 체크
        res_data['username'] = user.username        # mypage 정보
        res_data['email'] = user.email
        res_data['register'] = user.registerd_date
        res_data['userimg'] = fs.url(user.image)

        if res_data['userimg'] == "/media/":               # 이미지 체크
            res_data['img_check'] = 0
        else:
            res_data['img_check'] = 1

        if request.method == 'GET':
            return render(request, 'enter_exam1.html', res_data)
        elif request.method == 'POST':
            if user.check == False:
                return redirect('/main/enteroom/exam2')
            else:
                res_data['check'] = "차단이 완료되지 않았습니다."
                return render(request, 'enter_exam1.html', res_data)
    else:
        return redirect('/login')


@csrf_exempt
def exam2(request):
    res_data = {}
    fs = FileSystemStorage()
    user_session = request.session.get('user')
    if user_session:
        user = User.objects.get(pk=user_session)    # 로그인 체크
        res_data['username'] = user.username        # mypage 정보
        res_data['email'] = user.email
        res_data['register'] = user.registerd_date
        res_data['userimg'] = fs.url(user.image)

        if res_data['userimg'] == "/media/":               # 이미지 체크
            res_data['img_check'] = 0
        else:
            res_data['img_check'] = 1

        if request.method == 'GET':

            return render(request, 'enter_exam2.html', res_data)

        elif request.method == 'POST':
            print("POST")
            info = {}
            room_name = request.session.get('room_name')
            member_number = request.POST.get('member_number')
            member_name = request.POST.get('member_name')
            print("Get All DATA ")

            # 해당방의 DB속 명단Excel파일 조회
            room = Room.objects.get(room_name=room_name)
            member_file = room.file  # 명단
            print(member_file)

            # DB의 member_list로 회원번호 확인 및 index 추출
            member_list = room.member_list  # 회원번호만 적힌 리스트
            member_list = member_list[1:-1].split(', ')
            print(member_list)

            # CHECK NUMBER
            # Correct NUMBER
            if (member_number in member_list):
                # index=0은'회원번호(수험번호/학번)'이므로 index로 추출된 수 +1로 쓰면됨!
                member_index = member_list.index(member_number) + 1
                print('member_index:'+str(member_index))
                member = load_workbook("media/" + str(member_file))
                sheet = member['Sheet1']
                member_file_name = sheet['B'+str(member_index)].value

                # CHECK NAME
                # Wrong NAME
                if member_file_name != member_name:
                    info['result'] = "NO_NAME"
                    print('no_name')
                # Correct NAME
                else:
                    # WEB : 캡쳐이미지 받기
                    member_image_data = request.POST.__getitem__('photo')
                    member_image_data = member_image_data[22:]
                    member_image_path = str(
                        room_name)+'_'+str(member_number)+'_capture.png'
                    member_image = open(os.path.join(
                        FileSystemStorage().location)+str("/capture/")+member_image_path, "wb")
                    member_image.write(base64.b64decode(member_image_data))
                    member_image.close()

                    # exel 명단 속 이미지
                    image_loader = SheetImageLoader(sheet)
                    image = image_loader.get('C'+str(member_index))

                    member_file_image_path = (
                        room_name+"_"+str(member_number)+".jpg")
                    image.save("media/capture/"+member_file_image_path)
                    fs = FileSystemStorage()

                    # Face Recognition
                    a = (fs.location + str("/capture/") + member_file_image_path)
                    b = (fs.location + str("/capture/") + member_image_path)
                    print(a)
                    print(b)
                    # luxand API
                    luxand_client = luxand("12a42a8efedf4e24b84730ce440e5429")
                    member_file_image = luxand_client.add_person(
                        str(member_file_name), photos=[a])
                    result = luxand_client.verify(member_file_image, photo=b)
                    print(result)

                    # Recognition RESULT
                    if result['status'] == 'success':
                        info['result'] = "OK"
                        print("Recognition_SUCCESS")
                    else:
                        info['result'] = "NO_IMAGE_MATCH"
                        print("Recognition_FAIL")
            # Wrong NUMBER
            else:
                # 명단 속 존재하지 않는 회원번호 (입장불가!)
                info['result'] = "NO_MEMBER"
                print('no_member')
                # 해당페이지에서 팝업으로 입장불가표시주기
            return JsonResponse(info)
    else:
        return redirect('/login')


def exam3(request):
    res_data = {}
    fs = FileSystemStorage()
    user_session = request.session.get('user')
    if user_session:
        user = User.objects.get(pk=user_session)    # 로그인 체크
        res_data['username'] = user.username        # mypage 정보
        res_data['email'] = user.email
        res_data['register'] = user.registerd_date
        res_data['userimg'] = fs.url(user.image)

        if res_data['userimg'] == "/media/":               # 이미지 체크
            res_data['img_check'] = 0
        else:
            res_data['img_check'] = 1

        if request.method == 'GET':
            return render(request, 'enter_exam3.html', res_data)
        elif request.method == 'POST':
            room_session = request.session.get('room_name')
            room = Room.objects.get(room_name=room_session)
            roomname = room.room_name
            useremail = user.email
            roomowner = room.maker
            nickname = user.username
            roomtype = room.mode
            url = 'https://cranky-bohr-e0f18a.netlify.app/'+roomname + \
                '/'+useremail+'/'+roomowner+'/'+nickname+'/'+roomtype
            return redirect(url)
    else:
        return redirect('/login')


def study1(request):
    res_data = {}
    fs = FileSystemStorage()
    user_session = request.session.get('user')
    if user_session:
        user = User.objects.get(pk=user_session)    # 로그인 체크
        res_data['username'] = user.username        # mypage 정보
        res_data['email'] = user.email
        res_data['register'] = user.registerd_date
        res_data['userimg'] = fs.url(user.image)

        if res_data['userimg'] == "/media/":               # 이미지 체크
            res_data['img_check'] = 0
        else:
            res_data['img_check'] = 1

        if request.method == 'GET':
            return render(request, 'enter_study1.html', res_data)
        elif request.method == 'POST':
            if user.check == False:
                return redirect('/main/enteroom/study2')
            else:
                res_data['check'] = "차단이 완료되지 않았습니다."
                return render(request, 'enter_study1.html', res_data)
    else:
        return redirect('/login')


@method_decorator(csrf_exempt, name='dispatch')
def study2(request):
    res_data = {}
    fs = FileSystemStorage()
    user_session = request.session.get('user')
    res_data['session'] = user_session
    if user_session:
        user = User.objects.get(pk=user_session)    # 로그인 체크
        res_data['username'] = user.username        # mypage 정보
        res_data['email'] = user.email
        res_data['register'] = user.registerd_date
        res_data['userimg'] = fs.url(user.image)

        if res_data['userimg'] == "/media/":               # 이미지 체크
            res_data['img_check'] = 0
        else:
            res_data['img_check'] = 1

        if request.method == 'GET':
            return render(request, 'enter_study2.html', res_data)
        elif request.method == 'POST':
            room_session = request.session.get('room_name')
            room = Room.objects.get(room_name=room_session)
            roomname = room.room_name
            useremail = user.email
            roomowner = room.maker
            nickname = user.username
            roomtype = room.mode
            url = 'https://cranky-bohr-e0f18a.netlify.app/'+roomname + \
                '/'+useremail+'/'+roomowner+'/'+nickname+'/'+roomtype
            return redirect(url)
            # post_type = request.POST.get('enterRoom')
            # if post_type == 'toSsimong':   # 준영으로 넘어가는 나의 POST
            #     return  render(request,'ssimong.html',res_data)
            # else:   #  준영이 나에게 요청하는 POST

            #     # 준영에게 넘겨줄 data
            #     room_session = request.session.get('room_name')
            #     room = Room.objects.get(room_name=room_session)
            #     room_data = {}
            #     room_data['roomname'] = room.room_name
            #     room_data['useremail'] = user.email
            #     room_data['nickname'] = user.username
            #     room_data['roomowner'] = room.maker
            #     room_data['roomtype'] = room.mode
            #     return HttpResponse(simplejson.dumps(room_data))
    else:
        return redirect('/login')


def mylist(request):
    res_data = {}
    fs = FileSystemStorage()
    user_session = request.session.get('user')
    if user_session:
        user = User.objects.get(pk=user_session)    # 로그인 체크
        res_data['username'] = user.username        # mypage 정보
        res_data['email'] = user.email
        res_data['register'] = user.registerd_date
        res_data['userimg'] = fs.url(user.image)

        if res_data['userimg'] == "/media/":               # 이미지 체크
            res_data['img_check'] = 0
        else:
            res_data['img_check'] = 1

        if request.method == 'GET':
            return render(request, 'list.html', res_data)
        elif request.method == 'POST':
            return render(request, 'list,html', res_data)
    else:
        return redirect('/login')


# def room(request):
#     res_data={}
#     fs = FileSystemStorage()
#     user_session = request.session.get('user')
#     if user_session:
#         user = User.objects.get(pk=user_session)    # 로그인 체크
#         res_data['username'] = user.username        # mypage 정보
#         res_data['email'] = user.email
#         res_data['register'] = user.registerd_date
#         res_data['userimg'] = fs.url(user.image)

#         if res_data['userimg'] == "/media/":               # 이미지 체크
#             res_data['img_check'] = 0
#         else:
#             res_data['img_check'] = 1

#         if request.method == 'GET':
#             return render(request,'roomlist.html',res_data)
#         elif request.method == 'POST':
#             return  render(request,'roomlist,html',res_data)
#     else:
#         return redirect('/login')

class RoomList(ListView):
    model = Room
    template_name = 'roomlist.html'
    # context_object_name = "test"

    def get_queryset(self):    # roomlist를 보여줄 queryset 특정
        # session에 저장되어 있는 email과 room의 maker가 같은 것만 queryset에 넣음
        QuerySet = Room.objects.filter(
            maker=self.request.session.get('user_email')).order_by('-make_date')
        return QuerySet

    # def get_context_data(self):
    #     user = User.objects.get(pk= self.request.session.get('user'))
    #     res_data = {}
    #     res_data['username']= user.username
    #     return res_data


class AnalyticsList(ListView):
    model = Analytics
    template_name = 'analyticslist.html'

    def get_queryset(self):
        QuerySet = Analytics.objects.filter(
            email=self.request.session.get('user_email')).order_by('-make_date')
        return QuerySet


def analyticsDetail(request, pk):
    print("!!!!!!!!!!!!!!!!!!!!", pk)
    analytics = Analytics.objects.get(pk=pk)
    res_data = {}
    fs = FileSystemStorage()
    user_session = request.session.get('user')
    if user_session:
        user = User.objects.get(pk=user_session)    # 로그인 체크
        res_data['username'] = user.username        # mypage 정보
        res_data['email'] = user.email
        res_data['register'] = user.registerd_date
        res_data['userimg'] = fs.url(user.image)

        if res_data['userimg'] == "/media/":               # 이미지 체크
            res_data['img_check'] = 0
        else:
            res_data['img_check'] = 1

        # chartdata 선언
        dataSource = OrderedDict()
        dataSource["data"] = []  # chartdata는 json형식이다.
        dataSource["data"].append({"label": '앱 차단', "value": analytics.app})
        dataSource["data"].append({"label": '자리이탈', "value": analytics.person})
        dataSource["data"].append({"label": '학습 시간', "value": analytics.time})

        chartConfig = OrderedDict()
        chartConfig["caption"] = "집중도 통계"  # !!!!!!!!!!!!!!!!!!집중도 레벨 판별 해야함
        chartConfig["yAxisName"] = "점수"
        chartConfig["numberSuffix"] = "점"  # y축 숫자단위
        chartConfig["theme"] = "fusion"  # 테마

        dataSource["chart"] = chartConfig  # 그래프 특징 설정

        column2D = FusionCharts(
            "column2d", "myFirstChart", "500", "400", "chart-1", "json", dataSource)
        res_data['output'] = column2D.render()

        # res_data['count'] = analytics.count
        # res_data['rate'] = analytics.rate
        # res_data['level'] = analytics.level
        if request.method == 'GET':
            return render(request, 'list-analytics.html', res_data)
        elif request.method == 'POST':
            return render(request, 'list-analytics.html', res_data)
    else:
        return redirect('/login')


# !!!!!!!!!!!!!!!!!!!!!!!!!!!! 모든 단계가 끝나고 room_session 지워야 함
def analytics(request):
    res_data = {}
    fs = FileSystemStorage()
    user_session = request.session.get('user')
    if user_session:
        user = User.objects.get(pk=user_session)    # 로그인 체크
        res_data['username'] = user.username        # mypage 정보
        res_data['email'] = user.email
        res_data['register'] = user.registerd_date
        res_data['userimg'] = fs.url(user.image)

        if res_data['userimg'] == "/media/":               # 이미지 체크
            res_data['img_check'] = 0
        else:
            res_data['img_check'] = 1

        analytics = Analytics.objects.filter(email=user.email).last()

        # chartdata 선언
        dataSource = OrderedDict()
        dataSource["data"] = []  # chartdata는 json형식이다.
        dataSource["data"].append({"label": '앱 차단', "value": analytics.app})
        dataSource["data"].append({"label": '자리이탈', "value": analytics.person})
        dataSource["data"].append({"label": '학습 시간', "value": analytics.time})

        chartConfig = OrderedDict()
        chartConfig["caption"] = "집중도 통계"  # !!!!!!!!!!!!!!!!!!집중도 레벨 판별 해야함
        chartConfig["numberSuffix"] = "점"  # y축 숫자단위
        chartConfig["theme"] = "fusion"  # 테마

        dataSource["chart"] = chartConfig  # 그래프 특징 설정

        column2D = FusionCharts(
            "column2d", "myFirstChart", "500", "400", "chart-1", "json", dataSource)
        res_data['output'] = column2D.render()

        res_data['count'] = analytics.count
        res_data['rate'] = analytics.rate
        res_data['level'] = analytics.level

        if request.method == 'GET':
            return render(request, 'roomout-analytics.html', res_data)
        elif request.method == 'POST':  # 집중도에 사용할 데이터 받는 POST

            # !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!지우지 마세요
            # 1. 집중도 레벨
            # appPoint = 몇점 , personPoin = 몇점 , timePoint = 몇점
            # if(appPoint + pesonPoint + timePoint <= 299)
            #   level = 10
            # elif (appPoint + pesonPoint + timePoint <= 270 && appPoint + pesonPoint + timePoint < 299)
            #   level = 9
            # elif (appPoint + pesonPoint + timePoint <= 240 && appPoint + pesonPoint + timePoint < 270)
            #   level = 8
            # elif (appPoint + pesonPoint + timePoint <= 210 && appPoint + pesonPoint + timePoint < 240)
            #   level = 7
            # elif (appPoint + pesonPoint + timePoint <= 180 && appPoint + pesonPoint + timePoint < 210)
            #   level = 6
            # elif (appPoint + pesonPoint + timePoint <= 150 && appPoint + pesonPoint + timePoint < 180)
            #   level = 5
            # elif (appPoint + pesonPoint + timePoint <= 120 && appPoint + pesonPoint + timePoint < 150)
            #   level = 4
            # elif (appPoint + pesonPoint + timePoint <= 90 && appPoint + pesonPoint + timePoint < 120)
            #   level = 3
            # elif (appPoint + pesonPoint + timePoint <= 60 && appPoint + pesonPoint + timePoint < 90)
            #   level = 2
            # elif (appPoint + pesonPoint + timePoint < 60 )
            #   level = 1
            # else
            #   에러입니다.
            # room_name = request.session.get('room_name')
            # analytics = Analytics(room_name = room_name, email = user.email, grade = level, app = appPoint, person = personPoint = time = timePoint)
            # analytics.save()

            # 2. 등수  (모두 다 나왔을때, 나오지 않았을때 고려 해야함)
            # analytic = Analytics.object.filter('room_name = room_name').order_by('-level')  세션에 있는 room_name과 같은 통계 자료를 가져옴
            # num = 0
            # for x in analytic:
            #     if x.email == user.email:
            #         x.rate = num + 1
            #         x.save()
            #         print(x.rate)
            #     else:
            #         num = num+1

            # dataSource = OrderedDict()
            # dataSource["data"] = [] #chartdata는 json형식이다.
            # dataSource["data"].append({"label": '앱 차단', "value": appPoint})
            # dataSource["data"].append({"label": '자리이탈', "value": personPoint})
            # dataSource["data"].append({"label": '학습시간', "value": timePoint})

            # chartConfig = OrderedDict()
            # chartConfig["caption"] = "집중도 통계"
            # chartConfig["numberSuffix"] = "점" #y축 숫자단위
            # chartConfig["theme"] = "fusion" #테마

            # dataSource["chart"] = chartConfig # 그래프 특징 설정

            # column2D = FusionCharts("column2d", "myFirstChart", "500", "400", "chart-1", "json", dataSource)
            # res_data['output'] = column2D.render()

            # analytics = Analytics.object.fillter(email = user.name).last()
            # res_data['count'] = Analytics.object.filter(room_name = room_name).count()
            # analytic.count = res_data['count']
            # analytic.save()
            # res_data['rate'] = analytics.rate
            # res_data['level'] = analytics.level
            return render(request, 'roomout-analytics.html', res_data)
    else:
        return redirect('/login')


@csrf_exempt
def saveImages(request):
    data = request.POST.get('data')
    # data = data[22:]
    number = random.randrange(1, 10000)

    path = str(os.path.join(settings.STATIC_ROOT, 'resultImg/'))
    filename = 'image'+str(number) + '.png'

    image = open(path+filename, "wb")
    image.write(base64.b64decode(data))
    image.close()

    res_data = {}
    res_data['filename'] = filename
    return JsonResponse(res_data)


def roomout(request, time, mode):
    res_data = {}
    fs = FileSystemStorage()
    user_session = request.session.get('user')
    room_name_session = request.session.get('room_name')
    room = Room.objects.get(room_name=room_name_session)
    if user_session:
        user = User.objects.get(pk=user_session)    # 로그인 체크
        res_data['username'] = user.username        # mypage 정보
        res_data['email'] = user.email
        res_data['register'] = user.registerd_date
        res_data['userimg'] = fs.url(user.image)
        if res_data['userimg'] == "/media/":               # 이미지 체크
            res_data['img_check'] = 0
        else:
            res_data['img_check'] = 1

        print("!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!", room, room.maker)
        user = User.objects.get(email=room.maker)
        res_data['maker'] = user.username
        output = ''
        if mode == 'EXAM':
            if request.method == 'GET':
                return render(request, 'roomout.html', res_data)
            elif request.method == 'POST':
                # EXAM은 이제 끝났으니 room session 삭제
                del(request.session['room_name'])
                return render(request, 'roomout-success.html', res_data)
        elif mode == 'STUDY':
            time = time/1000/60
            if time >= 120:
                time = 100
            elif time < 120 and time >= 110:
                time = 95
            elif time < 110 and time >= 100:
                time = 90
            elif time < 100:
                time = time - 10
            else:
                time = 0

            analytics = Analytics.objects.filter(
                room_name=request.session.get('room_name'))
            if analytics:  # 해당 룸의 row가 있다.
                for x in analytics:
                    if(x.email == user.email):   # 해당 룸의 내 email을 가진 row가 있다.
                        output = 'YES'
                    else:
                        output = 'NO'
            elif not(analytics):  # 해당 룸의 row가 없다 -> 내가 제일 처음 -> 바로 생성
                analytics = Analytics(room_name=request.session.get(
                    'room_name'), email=user.email, time=time)
                analytics.save()

            if output == 'YES':  # 해당 룸의 row 중에 내 아이디의 row가 있다.
                analytics = Analytics.objects.filter(email=user.email).last()
                analytics.time = time
                analytics.save()
            elif output == 'NO':  # 해당 룸의 row 중에 내 아이디의 row가 없다.
                analytics = Analytics(room_name=request.session.get(
                    'room_name'), email=user.email, time=time)
                analytics.save()

            if request.method == 'GET':
                return render(request, 'roomout.html', res_data)
            elif request.method == 'POST':
                if user.check == False:
                    return redirect('/main/roomout/study')
                else:
                    res_data['check'] = "차단이 완료되지 않았습니다."
                    return render(request, 'roomout.html', res_data)
    else:
        return redirect('/login')


def roomoutExam(request):
    res_data = {}
    fs = FileSystemStorage()
    user_session = request.session.get('user')
    if user_session:
        user = User.objects.get(pk=user_session)    # 로그인 체크
        res_data['username'] = user.username        # mypage 정보
        res_data['email'] = user.email
        res_data['register'] = user.registerd_date
        res_data['userimg'] = fs.url(user.image)

        if res_data['userimg'] == "/media/":               # 이미지 체크
            res_data['img_check'] = 0
        else:
            res_data['img_check'] = 1

        if request.method == 'GET':
            return render(request, 'roomout.html', res_data)
        elif request.method == 'POST':
            if user.check == False:
                return redirect('/main/roomout/analytics')
            else:
                res_data['check'] = "차단이 완료되지 않았습니다."
                return render(request, 'roomout.html', res_data)
    else:
        return redirect('/login')


# web END

# android START

# 앱 이름 비번 명단 모두 제출시
@method_decorator(csrf_exempt, name='dispatch')
def app_makeroom(request):

    if request.method == "GET":
        return render(request, 'make_room.html')
    if request.method == "POST":
        print("post임")
        res_data = {}
        myfile = request.FILES['files']
        print(myfile)
        # fs = FileSystemStorage()
        # res_data['file_url'] = fs.url(myfile.name)
        myemail = request.POST.get('admin')
        myadmin = User.objects.get(email=myemail)
        myname = request.POST.get('name')
        mypass = request.POST.get('pass')
        mycheckbox = request.POST.get('checkbox')
        fs = FileSystemStorage()
        filename = fs.save(myfile.name, myfile)
        print(filename)

        member_list = []
        # myuser = Room(room_name=myname, room_password=mypass,
        #               maker=myadmin, mode=mycheckbox, member_list=member_list)
        # myuser.file = myfile
        # myuser.save()
        member = load_workbook("media/" + myfile.name)
        print(member)
        for cell in member['Sheet1']['A']:
            member_list.append(cell.value)
        print(member_list)
        myuser = Room(room_name=myname, room_password=mypass,
                      maker=myadmin, mode=mycheckbox, member_list=member_list)
        myuser.file = myfile
        os.remove(os.path.join(settings.MEDIA_ROOT, myfile.name))
        myuser.save()
        print(myemail)
        print(myname)
        print(mypass)
        print(mycheckbox)

        # test = json.loads(request.body)
        # myemail = test.get('admin')  # 아이디(이멜)

        # myname = test.get('name')
        # mypass = test.get('pass')
        # mycheckbox = test.get('checkbox')
        # print(myname)
        # #fs = FileSystemStorage()
        # #res_data['file_url'] = fs.url(file.name)
        # print(myfile)

        return HttpResponse(simplejson.dumps({"file": "ok"}))


# 앱 이름 비번만 제출시


@method_decorator(csrf_exempt, name='dispatch')
def app_makemyroom(request):
    if request.method == "GET":
        return render(request, 'home/make_room.html')
    if request.method == "POST":
        #res_data = {}
        myemail = request.POST.get('admin')  # 아이디(이멜)
        myadmin = User.objects.get(email=myemail)
        roomname = request.POST.get('roomname')
        password = request.POST.get('password')
        mycheckbox = request.POST.get('checkbox')
        member_list = []
        myuser = Room(room_name=roomname, room_password=password,
                      maker=myadmin, mode=mycheckbox, member_list=member_list)
        myuser.save()
        print(myuser)
        return HttpResponse(simplejson.dumps({"roomname": roomname, "password": password}))


@method_decorator(csrf_exempt, name='dispatch')
def app_enterroom(request):
    # 앱에서 오는 로그인 요청
    if request.method == "POST":
        roomname = request.POST.get('roomname', None)
        password = request.POST.get('password', None)
        # 받은 이메일이랑 비밀번호 =데이터와 일치하면
        # 리턴값으로 숫자 200 = 로그인 성공
        # 일치 안하면 숫자 100 = 로그인 실패
        if Room.objects.filter(room_name=roomname).exists():
            room = Room.objects.get(room_name=roomname)
            room_type = room.mode         # 룸 모드 가져오기
            # db에서 꺼내는 명령.Post로 받아온 username으로 , db의 username을 꺼내온다.
            print(password)
            print(room.room_password)
            if password == room.room_password:
                print(room.room_password)
                print(room_type)

                # 세션도 딕셔너리 변수 사용과 똑같이 사용하면 된다.
                # 세션 user라는 key에 방금 로그인한 id를 저장한것.
                # room_type(exam mode, study mode)구분해서 들고오기
                return HttpResponse(simplejson.dumps({"roomname": room_type}))
            else:
                # 방 비번 불일치
                return HttpResponse(simplejson.dumps({"roomname": "Fail"}))

        else:
            # 방 없음
            return HttpResponse(simplejson.dumps({"roomname": "None"}))


@method_decorator(csrf_exempt, name='dispatch')
def app_myroom(request):
    if request.method == "POST":
        email = request.POST.get('email', None)
        print(email)
        # 받은 이메일이랑 비밀번호 =데이터와 일치하면
        # 리턴값으로 숫자 200 = 로그인 성공
        # 일치 안하면 숫자 100 = 로그인 실패
        admin = User.objects.get(email=email)
        rooms = Room.objects.filter(maker=admin)

        # roomlist = []
        print(rooms)
        roomlist = serializers.serialize('json', rooms)

        return HttpResponse(roomlist, content_type="text/json-comment-filtered")


@method_decorator(csrf_exempt, name='dispatch')
def app_entermyroom(request):
    # 앱에서 오는 로그인 요청
    if request.method == "POST":
        roomname = request.POST.get('roomname', None)

        # 받은 이메일이랑 비밀번호 =데이터와 일치하면
        # 리턴값으로 숫자 200 = 로그인 성공
        # 일치 안하면 숫자 100 = 로그인 실패
        if Room.objects.filter(room_name=roomname).exists():
            room = Room.objects.get(room_name=roomname)
            room_type = room.mode        # 룸 모드 가져오기
            # db에서 꺼내는 명령.Post로 받아온 username으로 , db의 username을 꺼내온다.

            return HttpResponse(simplejson.dumps({"roomname": room_type}))


# EXAM 방

# 이미지 비교해야함

@method_decorator(csrf_exempt, name='dispatch')
def app_images(request):
    if request.method == "POST":
        # APP : 캡쳐이미지 받기
        capture_image = request.FILES['image']
        print(capture_image)
        # image.name = <room_name>_<member_number>_capture.png
        fs = FileSystemStorage()
        filename = fs.save("capture/"+capture_image.name, capture_image)
        uploaded_file_url = fs.url(filename)
        print(filename)
        print(uploaded_file_url)

        # image.name 에서 분리
        l = capture_image.name.split('_')
        room_name = l[0]
        member_index = l[1]

        print("Get All DATA ")

        # Room DB - excel 파일
        room = Room.objects.get(room_name=room_name)
        member_file = room.file  # 명단
        member = load_workbook("media/" + str(member_file))
        sheet = member['Sheet1']

        # exel 명단 속 이미지
        image_loader = SheetImageLoader(sheet)
        image = image_loader.get('C'+str(member_index))
        member_file_image_path = (room_name+"_"+str(member_index)+".jpg")
        image.save("media/capture/"+member_file_image_path)
        fs = FileSystemStorage()

        # Face Recognition
        a = (fs.location + str("\capture/") + member_file_image_path)
        b = (fs.location + str("\capture/") + capture_image.name)
        # luxand API
        luxand_client = luxand("12a42a8efedf4e24b84730ce440e5429")
        member_file_image = luxand_client.add_person(
            str(member_index), photos=[a])
        result = luxand_client.verify(member_file_image, photo=b)
        print(result)
        os.remove(os.path.join(settings.MEDIA_ROOT +
                               "/capture/", capture_image.name))
        # Recognition RESULT
        if result['status'] == 'success':
            print("Recognition_SUCCESS")
            return HttpResponse(simplejson.dumps({"image": "ok"}))  # 이미지 전송완료
        else:
            print("Recognition_FAIL")
            return HttpResponse(simplejson.dumps({"image": "no"}))  # 이미지 전송실패

# EXAM방 입장시 학번,이름 매칭확인


@method_decorator(csrf_exempt, name='dispatch')
def app_checkmyinfo(request):
    if request.method == "POST":
        info = {}
        # change_here
        room_name = request.POST.get('room', None)
        member_name = request.POST.get('name', None)
        member_number = request.POST.get('number', None)

        # room DB-member_list로 회원번호 확인 및 index 추출
        room = Room.objects.get(room_name=room_name)
        member_list = room.member_list  # 회원번호만 적힌 리스트
        member_list = member_list[1:-1].split(', ')
        print(member_list)

        # CHECK NUMBER
        # Correct NUMBER
        if (member_number in member_list):
            member_index = member_list.index(member_number) + 1
            print('member_index:'+str(member_index))

            # 해당방의 DB속 명단Excel파일 조회
            room = Room.objects.get(room_name=room_name)
            member_file = room.file  # 명단

            member = load_workbook("media/" + str(member_file))
            sheet = member['Sheet1']
            member_file_name = sheet['B'+str(member_index)].value

            # CHECK NAME
            # Wrong NAME
            if member_file_name != member_name:
                print('app_enterEXAM_info_no_match_name_num')
                return HttpResponse(simplejson.dumps({"roomname": "no",  "password": "no"}))
            # Correct NAME
            else:
                print('app_enterEXAM_info_success')
                return HttpResponse(simplejson.dumps({"roomname": "yes", "password": member_index}))

        # Wrong NUMBER
        else:
            print('app_enterEXAM_info_no_num')
            return HttpResponse(simplejson.dumps({"roomname": "fail", "password": "no"}))
